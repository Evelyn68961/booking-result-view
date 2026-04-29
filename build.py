"""Generate a self-contained HTML viewer + manager page for the leave-booking Excel file.

The manager-tab HTML is encrypted at build time with PBKDF2-SHA256 + AES-GCM keyed
on the manager password. Without the password, the manager UI markup is not present
in the DOM at all; viewing source only reveals an opaque base64 blob.
"""
import base64
import io
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).parent
XLSX = ROOT / "202401-202604預假紀錄.xlsx"
OUT = ROOT / "index.html"

# Manager password. Derives a 256-bit AES-GCM key via PBKDF2(SHA-256, 200_000 iters).
# To change: edit the plaintext below and rerun build.py.
MANAGER_PASSWORD = "FJUH.pharm0426"
PBKDF2_ITERS = 200_000


def encrypt_manager_block(plaintext: str, password: str) -> dict:
    salt = os.urandom(16)
    iv = os.urandom(12)
    key = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt,
                     iterations=PBKDF2_ITERS).derive(password.encode("utf-8"))
    ct = AESGCM(key).encrypt(iv, plaintext.encode("utf-8"), None)
    b64 = lambda b: base64.b64encode(b).decode("ascii")
    return {"salt": b64(salt), "iv": b64(iv), "ct": b64(ct), "iters": PBKDF2_ITERS}

EXCEL_EPOCH = datetime(1899, 12, 30)


def serial_to_iso(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        try:
            return (EXCEL_EPOCH + timedelta(days=float(value))).date().isoformat()
        except OverflowError:
            return str(value)
    return str(value)


def load_records():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    records = []
    for row in rows[1:]:
        if all(c is None or c == "" for c in row):
            continue
        rec = {h: v for h, v in zip(headers, row)}
        rec["_start_iso"] = serial_to_iso(rec.get("預假【起日】"))
        rec["_end_iso"] = serial_to_iso(rec.get("預假【迄日】"))
        records.append(rec)
    return headers, records


def to_jsonable(records):
    out = []
    for r in records:
        out.append({k: (v if not isinstance(v, datetime) else v.isoformat()) for k, v in r.items()})
    return out


HTML_TEMPLATE = r"""<!doctype html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>預假紀錄管理</title>
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
<style>
  :root {
    --bg: #f7f7fb;
    --panel: #ffffff;
    --border: #e3e6ee;
    --text: #1a1f36;
    --muted: #6b7393;
    --accent: #4f46e5;
    --pass: #16a34a;
    --fail: #dc2626;
    --pending: #d97706;
    --row-hover: #f1f3fb;
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; background: var(--bg); color: var(--text); font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft JhengHei", "PingFang TC", sans-serif; }
  header { padding: 16px 24px 0; background: var(--panel); border-bottom: 1px solid var(--border); }
  .head-row { display: flex; flex-wrap: wrap; gap: 12px; align-items: center; justify-content: space-between; }
  header h1 { font-size: 18px; margin: 0; font-weight: 600; }
  header .meta { color: var(--muted); font-size: 13px; }
  .tabs { display: flex; gap: 4px; margin-top: 14px; }
  .tab { padding: 10px 16px; background: transparent; border: none; border-bottom: 2px solid transparent; cursor: pointer; font-size: 14px; color: var(--muted); font-family: inherit; }
  .tab.active { color: var(--accent); border-bottom-color: var(--accent); font-weight: 600; }
  main { padding: 16px 24px 64px; }
  .panel { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px; margin-bottom: 14px; }
  .panel h2 { font-size: 14px; margin: 0 0 10px; font-weight: 600; color: #2d3656; }
  .controls { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; }
  .controls label { display: flex; flex-direction: column; gap: 4px; font-size: 12px; color: var(--muted); }
  .controls input, .controls select { padding: 8px 10px; border: 1px solid var(--border); border-radius: 6px; font-size: 14px; background: #fff; color: var(--text); font-family: inherit; }
  .controls input:focus, .controls select:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(79,70,229,0.15); }
  button.primary { background: var(--accent); color: white; border: none; padding: 8px 14px; border-radius: 6px; cursor: pointer; font-size: 14px; font-family: inherit; font-weight: 500; }
  button.primary:hover { background: #4338ca; }
  button.primary:disabled { opacity: 0.4; cursor: not-allowed; }
  button.ghost { background: transparent; color: var(--text); border: 1px solid var(--border); padding: 6px 12px; border-radius: 6px; cursor: pointer; font-size: 13px; font-family: inherit; }
  button.ghost:hover { background: var(--row-hover); }
  button.danger { background: transparent; color: var(--fail); border: 1px solid #fecaca; padding: 6px 12px; border-radius: 6px; cursor: pointer; font-size: 13px; font-family: inherit; }
  button.danger:hover { background: #fef2f2; }
  .summary { display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 12px; }
  .stat { background: var(--panel); border: 1px solid var(--border); border-radius: 8px; padding: 10px 14px; min-width: 120px; }
  .stat .n { font-size: 20px; font-weight: 600; }
  .stat .l { font-size: 12px; color: var(--muted); }
  .table-wrap { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; overflow: hidden; }
  table { width: 100%; border-collapse: collapse; font-size: 14px; }
  thead th { position: sticky; top: 0; background: #f0f2f8; padding: 10px 12px; text-align: left; font-weight: 600; font-size: 13px; color: #2d3656; cursor: pointer; user-select: none; border-bottom: 1px solid var(--border); white-space: nowrap; }
  thead th:hover { background: #e6eaf4; }
  thead th.sorted-asc::after { content: " ▲"; color: var(--accent); font-size: 11px; }
  thead th.sorted-desc::after { content: " ▼"; color: var(--accent); font-size: 11px; }
  thead th.no-sort { cursor: default; }
  thead th.no-sort:hover { background: #f0f2f8; }
  tbody td { padding: 8px 12px; border-bottom: 1px solid #eef0f6; vertical-align: top; }
  tbody tr:hover { background: var(--row-hover); }
  .pill { display: inline-block; padding: 2px 8px; border-radius: 999px; font-size: 12px; font-weight: 500; white-space: nowrap; }
  .pill.pass { background: #dcfce7; color: var(--pass); }
  .pill.fail { background: #fee2e2; color: var(--fail); }
  .pill.pending { background: #fef3c7; color: var(--pending); }
  .pill.other { background: #e5e7eb; color: #4b5563; }
  .pill.batch { background: #ede9fe; color: var(--accent); }
  .empty { padding: 40px; text-align: center; color: var(--muted); }
  .pager { display: flex; gap: 8px; justify-content: flex-end; align-items: center; padding: 10px 14px; border-top: 1px solid var(--border); background: #fafbff; font-size: 13px; color: var(--muted); }
  .pager button { padding: 4px 10px; border: 1px solid var(--border); background: #fff; border-radius: 6px; cursor: pointer; font-size: 13px; }
  .pager button:disabled { opacity: 0.4; cursor: not-allowed; }
  .reason { color: var(--muted); font-size: 12px; }
  td.name { font-weight: 600; white-space: nowrap; }
  td.date { white-space: nowrap; font-variant-numeric: tabular-nums; color: #2d3656; }
  td.num  { text-align: right; font-variant-numeric: tabular-nums; }
  .drop { border: 2px dashed var(--border); border-radius: 10px; padding: 28px; text-align: center; color: var(--muted); cursor: pointer; transition: all 0.15s; }
  .drop.dragover { border-color: var(--accent); background: #eef0ff; color: var(--accent); }
  .drop strong { color: var(--accent); }
  .help { font-size: 12px; color: var(--muted); margin-top: 6px; }
  .actions-row { display: flex; gap: 8px; flex-wrap: wrap; align-items: center; margin-top: 12px; }
  .actions-row .spacer { flex: 1; }
  .conflict-list { font-size: 11px; color: var(--fail); margin-top: 4px; line-height: 1.4; }
  .ok-list { font-size: 11px; color: var(--pass); margin-top: 4px; line-height: 1.4; }
  .toast { position: fixed; bottom: 24px; right: 24px; background: #1a1f36; color: white; padding: 10px 16px; border-radius: 8px; font-size: 13px; opacity: 0; pointer-events: none; transition: opacity 0.2s; z-index: 1000; }
  .toast.show { opacity: 1; }
  select.decision { padding: 4px 6px; border: 1px solid var(--border); border-radius: 4px; font-size: 12px; background: #fff; font-family: inherit; }
  input.edit-cell, select.edit-cell { padding: 4px 6px; border: 1px solid var(--border); border-radius: 4px; font-size: 12px; background: #fff; font-family: inherit; width: 100%; min-width: 80px; }
  input.edit-cell:focus, select.edit-cell:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 2px rgba(79,70,229,0.15); }
  input.edit-cell.edit-name { font-weight: 600; min-width: 80px; }
  input.edit-cell.edit-date { min-width: 130px; font-variant-numeric: tabular-nums; }
  .modal-overlay { position: fixed; inset: 0; background: rgba(15,18,30,0.45); display: flex; align-items: center; justify-content: center; z-index: 1100; }
  .modal-overlay.hidden { display: none !important; }
  .modal-card { background: var(--panel); border-radius: 12px; padding: 22px; max-width: 460px; width: 92%; box-shadow: 0 10px 30px rgba(0,0,0,0.18); }
  .modal-card h3 { margin: 0 0 6px; font-size: 16px; }
  .modal-card .modal-meta { color: var(--muted); font-size: 12px; margin-bottom: 14px; }
  .modal-fields { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }
  .modal-fields label { display: flex; flex-direction: column; gap: 4px; font-size: 12px; color: var(--muted); }
  .modal-fields label.wide { grid-column: span 2; }
  .modal-fields input, .modal-fields select { padding: 8px 10px; border: 1px solid var(--border); border-radius: 6px; font-size: 14px; font-family: inherit; }
  .modal-fields input:focus, .modal-fields select:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(79,70,229,0.15); }
  .modal-actions { display: flex; gap: 8px; justify-content: flex-end; margin-top: 18px; }
  td.actions { white-space: nowrap; }
  td.actions button { margin-right: 4px; }
  details.day-detail { margin-top: 10px; }
  details.day-detail summary { cursor: pointer; font-size: 13px; color: var(--accent); padding: 6px 0; }
  .day-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(140px, 1fr)); gap: 6px; margin-top: 8px; font-size: 12px; }
  .day-cell { padding: 6px 8px; border-radius: 6px; background: #f1f3fb; }
  .day-cell.full { background: #fee2e2; color: var(--fail); }
  .day-cell.tight { background: #fef3c7; color: var(--pending); }
  .day-cell .d { font-variant-numeric: tabular-nums; }
  .day-cell .n { font-weight: 600; }
  .cal-toolbar { display: flex; gap: 12px; align-items: center; margin-bottom: 14px; flex-wrap: wrap; }
  .cal-toolbar button { padding: 6px 12px; background: var(--panel); border: 1px solid var(--border); border-radius: 6px; cursor: pointer; font-family: inherit; font-size: 13px; }
  .cal-toolbar button:hover { background: #f5f6fa; }
  .cal-month-label { font-size: 16px; font-weight: 600; min-width: 130px; text-align: center; }
  .cal-legend { display: flex; gap: 14px; margin-left: auto; font-size: 12px; color: var(--muted); align-items: center; }
  .cal-legend .swatch { display: inline-block; width: 14px; height: 14px; border-radius: 3px; margin-right: 4px; vertical-align: middle; border: 1px solid var(--border); }
  .cal-grid { display: grid; grid-template-columns: repeat(7, 1fr); gap: 1px; background: var(--border); border: 1px solid var(--border); border-radius: 6px; overflow: hidden; }
  .cal-head { background: #f5f6fa; padding: 8px; text-align: center; font-size: 12px; color: var(--muted); font-weight: 600; }
  .cal-cell { background: var(--panel); padding: 8px 6px; min-height: 60px; cursor: pointer; user-select: none; transition: filter 0.15s; }
  .cal-cell:hover:not(.empty) { filter: brightness(0.94); }
  .cal-cell.empty { background: #fafbfd; color: #cbd0db; cursor: default; }
  .cal-cell .num { font-size: 13px; font-variant-numeric: tabular-nums; font-weight: 500; }
  .cal-cell .cnt { font-size: 11px; margin-top: 6px; color: var(--muted); }
  .cal-cell.lvl0 { background: #f0fdf4; }
  .cal-cell.lvl1 { background: #fef9c3; }
  .cal-cell.lvl2 { background: #fecaca; }
  .cal-cell.lvl0 .cnt { color: #16a34a; }
  .cal-cell.lvl1 .cnt { color: #ca8a04; }
  .cal-cell.lvl2 .cnt { color: #dc2626; font-weight: 600; }
  .cal-cell.selected { outline: 2px solid var(--accent); outline-offset: -2px; z-index: 1; position: relative; }
  .cal-day-list { width: 100%; font-size: 13px; border-collapse: collapse; }
  .cal-day-list th, .cal-day-list td { text-align: left; padding: 6px 8px; border-bottom: 1px solid var(--border); }
  .cal-day-list th { font-weight: 600; color: var(--muted); font-size: 12px; }
  .hidden { display: none !important; }
  .lock-card { max-width: 420px; margin: 60px auto; padding: 32px; background: var(--panel); border: 1px solid var(--border); border-radius: 12px; text-align: center; }
  .lock-card .icon { font-size: 32px; margin-bottom: 8px; }
  .lock-card h2 { font-size: 18px; margin: 0 0 6px; }
  .lock-card p { font-size: 13px; color: var(--muted); margin: 0 0 20px; }
  .lock-card form { display: flex; gap: 8px; justify-content: center; }
  .lock-card input { flex: 1; padding: 10px 12px; border: 1px solid var(--border); border-radius: 6px; font-size: 14px; font-family: inherit; }
  .lock-card input:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(79,70,229,0.15); }
  .lock-card .err { color: var(--fail); font-size: 12px; margin-top: 12px; min-height: 16px; }
  .logout-bar { display: flex; justify-content: flex-end; margin-bottom: 10px; }
</style>
</head>
<body>
<header>
  <div class="head-row">
    <div>
      <h1>預假紀錄管理系統</h1>
      <div class="meta">資料來源：202401-202604預假紀錄.xlsx</div>
    </div>
    <div class="meta" id="rangeMeta"></div>
  </div>
  <div class="tabs">
    <button class="tab active" data-tab="view">檢視紀錄</button>
    <button class="tab" data-tab="calendar">預約日曆</button>
    <button class="tab" data-tab="manager">新申請審核（管理員）</button>
  </div>
</header>

<main>
  <!-- ============ VIEW TAB ============ -->
  <section id="tab-view">
    <div class="panel">
      <div class="controls">
        <label>關鍵字搜尋
          <input id="q" type="search" placeholder="姓名、結果、原因…" />
        </label>
        <label>姓名
          <select id="fName"><option value="">全部</option></select>
        </label>
        <label>審核結果
          <select id="fStatus">
            <option value="">全部</option>
            <option value="pass">通過</option>
            <option value="fail">未通過</option>
            <option value="other">其他</option>
          </select>
        </label>
        <label>未通過原因
          <select id="fReason"><option value="">全部</option></select>
        </label>
        <label>起日 從
          <input id="fFrom" type="date" />
        </label>
        <label>起日 到
          <input id="fTo" type="date" />
        </label>
        <label>每頁筆數
          <select id="pageSize">
            <option>25</option><option selected>50</option><option>100</option><option>250</option><option value="0">全部</option>
          </select>
        </label>
      </div>
    </div>

    <div class="summary" id="summary"></div>

    <div class="table-wrap">
      <table id="t">
        <thead><tr id="head"></tr></thead>
        <tbody id="body"></tbody>
      </table>
      <div class="pager" id="pager"></div>
    </div>
  </section>

  <!-- ============ CALENDAR TAB ============ -->
  <section id="tab-calendar" class="hidden">
    <div class="panel">
      <div class="cal-toolbar">
        <button id="calPrev">‹ 上個月</button>
        <span class="cal-month-label" id="calMonthLabel"></span>
        <button id="calNext">下個月 ›</button>
        <button id="calToday">回到今天</button>
        <span class="cal-legend">
          <span><span class="swatch" style="background:#f0fdf4;"></span>空</span>
          <span><span class="swatch" style="background:#fef9c3;"></span>1 人</span>
          <span><span class="swatch" style="background:#fecaca;"></span>2 人以上</span>
        </span>
      </div>
      <div class="cal-grid" id="calGrid"></div>
    </div>
    <div class="panel">
      <h2 id="calDayTitle">點選日期以查看當日預假人員</h2>
      <div id="calDayBody" class="help">尚未選擇日期。</div>
    </div>
  </section>

  <!-- ============ MANAGER TAB ============ -->
  <section id="tab-manager" class="hidden">
    <div id="lockScreen" class="lock-card">
      <div class="icon">🔒</div>
      <h2>管理員區域</h2>
      <p>請輸入密碼以使用新申請審核功能</p>
      <form id="unlockForm" autocomplete="off">
        <input id="pwInput" type="password" autocomplete="new-password" placeholder="密碼" />
        <button class="primary" type="submit">解鎖</button>
      </form>
      <div class="err" id="unlockErr"></div>
    </div>
    <div id="managerContent"></div>
  </section>
</main>

<div class="toast" id="toast"></div>

<div id="editModal" class="modal-overlay hidden">
  <div class="modal-card">
    <h3>編輯紀錄</h3>
    <div class="modal-meta" id="emMeta"></div>
    <div class="modal-fields">
      <label class="wide">姓名 <input id="emName" type="text" /></label>
      <label>結果
        <select id="emStatus">
          <option value="pass">通過</option>
          <option value="fail">未通過</option>
        </select>
      </label>
      <label>類別
        <select id="emCat">
          <option value="">未填</option>
          <option value="1-3天">1-3天</option>
          <option value="4-10天">4-10天</option>
          <option value=">10天">10天以上</option>
        </select>
      </label>
      <label class="wide">拒絕原因 <input id="emReason" type="text" placeholder="(僅未通過時使用)" /></label>
      <label>起日 <input id="emStart" type="date" /></label>
      <label>迄日 <input id="emEnd" type="date" /></label>
    </div>
    <div class="modal-actions">
      <button class="ghost" id="emCancel">取消</button>
      <button class="primary" id="emSave">儲存</button>
    </div>
  </div>
</div>

<script>
const BAKED = __DATA__;
const HEADERS = __HEADERS__;
const STORAGE_KEY = 'booking-extra-records-v1';
const BATCH_KEY = 'booking-batch-v1';
const BAKED_PATCHES_KEY = 'booking-baked-patches-v1';

// =============== STATE ===============
const state = {
  q: '', name: '', status: '', reason: '',
  from: '', to: '',
  sortKey: '_start_iso', sortDir: 'desc',
  page: 1, pageSize: 50,
  quota: 2, minDays: 4, maxDays: 10, yearlyPoints: 12,
  gateDay: '',
  batch: [],
};

const $ = (id) => document.getElementById(id);

// =============== STORAGE ===============
function loadStored() {
  let arr;
  try { arr = JSON.parse(localStorage.getItem(STORAGE_KEY) || '[]'); }
  catch { arr = []; }
  let dirty = false;
  for (const r of arr) { if (!r._id) { r._id = uid(); dirty = true; } }
  if (dirty) localStorage.setItem(STORAGE_KEY, JSON.stringify(arr));
  return arr;
}
function saveStored(arr) { localStorage.setItem(STORAGE_KEY, JSON.stringify(arr)); }
function loadBatch() {
  try { return JSON.parse(localStorage.getItem(BATCH_KEY) || '[]'); }
  catch { return []; }
}
function saveBatch() { localStorage.setItem(BATCH_KEY, JSON.stringify(state.batch)); }
function loadBakedPatches() {
  try { return JSON.parse(localStorage.getItem(BAKED_PATCHES_KEY) || '{}'); }
  catch { return {}; }
}
function saveBakedPatches(p) { localStorage.setItem(BAKED_PATCHES_KEY, JSON.stringify(p)); }

// Build effective baked records: BAKED merged with patches (edits/deletes) from localStorage.
// Each row carries `_baked_idx` so the UI can target the correct slot for further edits.
function effectiveBaked() {
  const patches = loadBakedPatches();
  const out = [];
  for (let i = 0; i < BAKED.length; i++) {
    const p = patches[i];
    if (p && p._deleted) continue;
    out.push(p ? Object.assign({}, BAKED[i], p, { _baked_idx: i }) : Object.assign({}, BAKED[i], { _baked_idx: i }));
  }
  return out;
}

// All records (effective baked + manager-committed) used for both views and quota math.
function allRecords() { return effectiveBaked().concat(loadStored()); }

// =============== HELPERS ===============
function classifyStatus(s) {
  if (!s) return 'other';
  const t = String(s);
  if (t.startsWith('通過')) return 'pass';
  if (t.startsWith('未通過')) return 'fail';
  if (t.includes('待') || t.includes('審核中')) return 'pending';
  return 'other';
}
function reasonOf(s) {
  if (!s) return '';
  const m = String(s).match(/未通過\s*[-–—]\s*(.+)/);
  return m ? m[1].trim() : '';
}
function escapeHtml(s) {
  return String(s ?? '').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
}
function uniqueSorted(values) {
  return Array.from(new Set(values.filter(v => v !== '' && v != null))).sort((a, b) => String(a).localeCompare(String(b), 'zh-Hant'));
}
function toast(msg) {
  const el = $('toast'); el.textContent = msg; el.classList.add('show');
  clearTimeout(toast._t); toast._t = setTimeout(() => el.classList.remove('show'), 2200);
}
function uid() { return Math.random().toString(36).slice(2, 10); }
function daysInRange(a, b) {
  // a, b: ISO date strings; inclusive count
  const ad = new Date(a + 'T00:00:00'), bd = new Date(b + 'T00:00:00');
  return Math.round((bd - ad) / 86400000) + 1;
}
function* iterDates(a, b) {
  const ad = new Date(a + 'T00:00:00'), bd = new Date(b + 'T00:00:00');
  for (let d = ad; d <= bd; d.setDate(d.getDate() + 1)) {
    yield d.toISOString().slice(0, 10);
  }
}
function isWeekend(iso) {
  const dow = new Date(iso + 'T00:00:00').getDay();
  return dow === 0 || dow === 6;
}
// Compute round window end: first Sunday of the month after Gate Day + 7 months (inclusive).
function roundWindow() {
  if (!state.gateDay) return null;
  const d = new Date(state.gateDay + 'T00:00:00');
  // Day-1 of (Gate Day's month + 8) — handles year rollover automatically.
  const end = new Date(d.getFullYear(), d.getMonth() + 8, 1);
  // Advance to the Sunday on or after the 1st. JS getDay(): Sun=0, Mon=1 ... Sat=6.
  const dow = end.getDay();
  end.setDate(end.getDate() + ((7 - dow) % 7));
  const iso = (x) => new Date(x.getTime() - x.getTimezoneOffset() * 60000).toISOString().slice(0, 10);
  return { from: state.gateDay, to: iso(end) };
}
function inRoundWindow(iso) {
  const w = roundWindow(); if (!w) return true;
  return iso >= w.from && iso <= w.to;
}

// Convert various inputs to YYYY-MM-DD ISO. Accepts Excel serial, Date object, Date strings.
const EPOCH = Date.UTC(1899, 11, 30);
function toIso(v) {
  if (v == null || v === '') return null;
  if (v instanceof Date) return new Date(v.getTime() - v.getTimezoneOffset()*60000).toISOString().slice(0, 10);
  if (typeof v === 'number') {
    const ms = EPOCH + v * 86400000;
    return new Date(ms).toISOString().slice(0, 10);
  }
  const s = String(v).trim();
  // Already ISO?
  let m = s.match(/^(\d{4})-(\d{1,2})-(\d{1,2})/);
  if (m) return `${m[1]}-${m[2].padStart(2,'0')}-${m[3].padStart(2,'0')}`;
  // YYYY/M/D
  m = s.match(/^(\d{4})[\/.](\d{1,2})[\/.](\d{1,2})/);
  if (m) return `${m[1]}-${m[2].padStart(2,'0')}-${m[3].padStart(2,'0')}`;
  // 2026年4月8日
  m = s.match(/(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日/);
  if (m) return `${m[1]}-${m[2].padStart(2,'0')}-${m[3].padStart(2,'0')}`;
  // Numeric serial in string?
  if (/^\d+(\.\d+)?$/.test(s)) return toIso(Number(s));
  // Last resort
  const d = new Date(s);
  if (!isNaN(d)) return new Date(d.getTime() - d.getTimezoneOffset()*60000).toISOString().slice(0, 10);
  return null;
}

// =============== APPROVED-DAY MAP ===============
// Build a map { isoDate -> approvedCount } from baked + stored + already-approved batch entries.
function buildDayMap(includeBatchApproved = true) {
  const map = new Map();
  function add(start, end) {
    if (!start || !end) return;
    if (end < start) return;
    for (const d of iterDates(start, end)) {
      map.set(d, (map.get(d) || 0) + 1);
    }
  }
  for (const r of allRecords()) {
    if (classifyStatus(r['審核結果']) !== 'pass') continue;
    add(r._start_iso, r._end_iso);
  }
  if (includeBatchApproved) {
    for (const b of state.batch) {
      if (b.decision === 'pass') add(b.start, b.end);
    }
  }
  return map;
}

// Approved-submission count for one person in a calendar year (history only; baked + stored).
function personPointsUsed(name, year) {
  let n = 0;
  for (const r of allRecords()) {
    if (r['你的名字'] !== name) continue;
    if (classifyStatus(r['審核結果']) !== 'pass') continue;
    if (!r._start_iso) continue;
    if (Number(r._start_iso.slice(0, 4)) === year) n++;
  }
  return n;
}

// Predict result for one batch entry. `personYearUsed` is keyed by `${name}|${year}` and holds
// the running tally including earlier-in-batch approved entries.
function predict(entry, dayMap, personYearUsed) {
  if (!entry.name) return { cls: 'fail', reason: '缺少姓名', conflicts: [] };
  if (!entry.start || !entry.end) return { cls: 'fail', reason: '缺少日期', conflicts: [] };
  if (entry.end < entry.start) return { cls: 'fail', reason: '預假天數錯誤（迄日早於起日）', conflicts: [] };

  const days = daysInRange(entry.start, entry.end);
  if (days < state.minDays || days > state.maxDays) {
    return { cls: 'fail', reason: `預假天數錯誤（${days} 天不在 ${state.minDays}–${state.maxDays} 範圍）`, conflicts: [] };
  }

  const w = roundWindow();
  if (w && (entry.start < w.from || entry.end > w.to)) {
    return { cls: 'fail', reason: `超出可預約範圍（${w.from} ～ ${w.to}）`, conflicts: [] };
  }

  const year = Number(entry.start.slice(0, 4));
  const key  = `${entry.name}|${year}`;
  const used = personYearUsed.has(key) ? personYearUsed.get(key) : personPointsUsed(entry.name, year);
  if (used + 1 > state.yearlyPoints) {
    return { cls: 'fail', reason: `年度點數不足（${year} 年已核准 ${used} 次，上限 ${state.yearlyPoints}）`, conflicts: [] };
  }

  const conflicts = [];
  for (const d of iterDates(entry.start, entry.end)) {
    const cur = dayMap.get(d) || 0;
    if (cur >= state.quota) conflicts.push({ d, cur });
  }
  if (conflicts.length) {
    return { cls: 'fail', reason: '已超過上限人數', conflicts };
  }
  return { cls: 'pass', reason: '', conflicts: [] };
}

// Re-run prediction for every batch entry. Earlier "pass" entries occupy quota slots and
// consume yearly points before later entries are tested.
function recomputeBatch() {
  const baseMap = buildDayMap(false); // history only
  const personYearUsed = new Map();
  for (const e of state.batch) {
    if (!e.name || !e.start) continue;
    const year = Number(e.start.slice(0, 4));
    const key = `${e.name}|${year}`;
    if (!personYearUsed.has(key)) personYearUsed.set(key, personPointsUsed(e.name, year));
  }
  for (const e of state.batch) {
    e.predicted = predict(e, baseMap, personYearUsed);
    const eff = e.decision === 'auto' ? e.predicted.cls : e.decision;
    if (eff === 'pass' && e.start && e.end && e.end >= e.start) {
      for (const d of iterDates(e.start, e.end)) {
        baseMap.set(d, (baseMap.get(d) || 0) + 1);
      }
      const key = `${e.name}|${Number(e.start.slice(0, 4))}`;
      personYearUsed.set(key, (personYearUsed.get(key) || 0) + 1);
    }
  }
}

// =============== VIEW TAB ===============
const COLS = [
  { key: '你的名字',       label: '姓名', cls: 'name' },
  { key: '審核結果',       label: '審核結果', cls: 'status' },
  { key: '_start_iso',    label: '預假起日', cls: 'date', sortAs: 'date' },
  { key: '_end_iso',      label: '預假迄日', cls: 'date', sortAs: 'date' },
  { key: '預假天數',       label: '天數區間' },
  { key: '送出時間',       label: '送出時間', cls: 'date' },
  { key: '特休',          label: '特休', cls: 'num', sortAs: 'num' },
  { key: '時數',          label: '時數', cls: 'num', sortAs: 'num' },
];

function applyFilters() {
  const q = state.q.trim().toLowerCase();
  return allRecords().filter(r => {
    if (state.name && r['你的名字'] !== state.name) return false;
    const cls = classifyStatus(r['審核結果']);
    if (state.status && cls !== state.status) return false;
    if (state.reason && reasonOf(r['審核結果']) !== state.reason) return false;
    if (state.from && (!r._start_iso || r._start_iso < state.from)) return false;
    if (state.to && (!r._start_iso || r._start_iso > state.to)) return false;
    if (q) {
      const hay = [r['你的名字'], r['審核結果'], r['送出時間'], r['預假天數'], r._start_iso, r._end_iso, r['特休'], r['時數']]
        .map(v => String(v ?? '').toLowerCase()).join(' ');
      if (!hay.includes(q)) return false;
    }
    return true;
  });
}

function sortRows(rows) {
  const col = COLS.find(c => c.key === state.sortKey);
  const dir = state.sortDir === 'asc' ? 1 : -1;
  const sortAs = col?.sortAs;
  return rows.slice().sort((a, b) => {
    let va = a[state.sortKey], vb = b[state.sortKey];
    if (sortAs === 'num') { va = Number(va); vb = Number(vb); if (Number.isNaN(va)) va = -Infinity; if (Number.isNaN(vb)) vb = -Infinity; }
    else if (sortAs === 'date') { va = va || ''; vb = vb || ''; }
    else { va = String(va ?? ''); vb = String(vb ?? ''); }
    if (va < vb) return -1 * dir;
    if (va > vb) return  1 * dir;
    return 0;
  });
}

function renderHead() {
  const tr = $('head');
  const cells = COLS.map(c => {
    const sortedCls = state.sortKey === c.key ? (state.sortDir === 'asc' ? 'sorted-asc' : 'sorted-desc') : '';
    return `<th class="${sortedCls}" data-key="${c.key}">${escapeHtml(c.label)}</th>`;
  });
  if (MANAGER_UNLOCKED) cells.push('<th class="no-sort">操作</th>');
  tr.innerHTML = cells.join('');
  tr.querySelectorAll('th[data-key]').forEach(th => {
    th.onclick = () => {
      const k = th.dataset.key;
      if (state.sortKey === k) state.sortDir = state.sortDir === 'asc' ? 'desc' : 'asc';
      else { state.sortKey = k; state.sortDir = 'asc'; }
      renderView();
    };
  });
}

function recordHandle(r) {
  if (r._source === 'manager' && r._id) return `m:${r._id}`;
  if (r._baked_idx !== undefined) return `b:${r._baked_idx}`;
  return '';
}

function renderBody(rows) {
  const body = $('body');
  const colCount = COLS.length + (MANAGER_UNLOCKED ? 1 : 0);
  if (!rows.length) { body.innerHTML = `<tr><td colspan="${colCount}" class="empty">沒有符合條件的紀錄</td></tr>`; return; }
  body.innerHTML = rows.map(r => {
    const cells = COLS.map(c => {
      let v = r[c.key];
      if (c.key === '審核結果') {
        const cls = classifyStatus(v);
        const reason = reasonOf(v);
        const label = cls === 'pass' ? '通過' : cls === 'fail' ? '未通過' : (v ?? '');
        const pill = `<span class="pill ${cls}">${escapeHtml(label)}</span>`;
        const tag = r._source === 'manager' ? ' <span class="pill batch">手動加入</span>' : '';
        return `<td>${pill}${tag}${reason ? `<div class="reason">${escapeHtml(reason)}</div>` : ''}</td>`;
      }
      if (v === null || v === undefined || v === '') v = '';
      return `<td class="${c.cls || ''}">${escapeHtml(v)}</td>`;
    });
    if (MANAGER_UNLOCKED) {
      const h = recordHandle(r);
      cells.push(`<td class="actions"><button class="ghost" data-edit="${h}">編輯</button><button class="danger" data-del="${h}">刪除</button></td>`);
    }
    return '<tr>' + cells.join('') + '</tr>';
  }).join('');

  if (MANAGER_UNLOCKED) {
    body.querySelectorAll('button[data-edit]').forEach(b => b.onclick = () => openEditModal(b.dataset.edit));
    body.querySelectorAll('button[data-del]').forEach(b => b.onclick = () => deleteRecordByHandle(b.dataset.del));
  }
}

function renderSummary(filtered) {
  const total = filtered.length;
  const pass = filtered.filter(r => classifyStatus(r['審核結果']) === 'pass').length;
  const fail = filtered.filter(r => classifyStatus(r['審核結果']) === 'fail').length;
  const other = total - pass - fail;
  const uniqueNames = new Set(filtered.map(r => r['你的名字'])).size;
  $('summary').innerHTML = `
    <div class="stat"><div class="n">${total}</div><div class="l">符合筆數 / 共 ${allRecords().length}</div></div>
    <div class="stat"><div class="n" style="color:var(--pass)">${pass}</div><div class="l">通過</div></div>
    <div class="stat"><div class="n" style="color:var(--fail)">${fail}</div><div class="l">未通過</div></div>
    <div class="stat"><div class="n">${other}</div><div class="l">其他</div></div>
    <div class="stat"><div class="n">${uniqueNames}</div><div class="l">不重複姓名</div></div>
  `;
}

function renderPager(totalRows) {
  const ps = state.pageSize;
  if (ps === 0) { $('pager').innerHTML = `共 ${totalRows} 筆`; return; }
  const totalPages = Math.max(1, Math.ceil(totalRows / ps));
  if (state.page > totalPages) state.page = totalPages;
  const from = totalRows === 0 ? 0 : (state.page - 1) * ps + 1;
  const to = Math.min(totalRows, state.page * ps);
  $('pager').innerHTML = `
    <span>顯示 ${from}–${to} / 共 ${totalRows} 筆</span>
    <button id="pPrev" ${state.page <= 1 ? 'disabled' : ''}>上一頁</button>
    <span>第 ${state.page} / ${totalPages} 頁</span>
    <button id="pNext" ${state.page >= totalPages ? 'disabled' : ''}>下一頁</button>
  `;
  const prev = document.getElementById('pPrev'); if (prev) prev.onclick = () => { state.page--; renderView(); };
  const next = document.getElementById('pNext'); if (next) next.onclick = () => { state.page++; renderView(); };
}

function paginate(rows) {
  if (state.pageSize === 0) return rows;
  const start = (state.page - 1) * state.pageSize;
  return rows.slice(start, start + state.pageSize);
}

function renderView() {
  refreshFilters();
  const filtered = applyFilters();
  const sorted = sortRows(filtered);
  renderHead();
  renderSummary(filtered);
  renderBody(paginate(sorted));
  renderPager(filtered.length);
}

function refreshFilters() {
  const all = allRecords();
  const names = uniqueSorted(all.map(r => r['你的名字']));
  const reasons = uniqueSorted(all.map(r => reasonOf(r['審核結果'])));
  const fName = $('fName');
  const fReason = $('fReason');
  const prevName = fName.value, prevReason = fReason.value;
  fName.innerHTML = '<option value="">全部</option>' + names.map(n => `<option>${escapeHtml(n)}</option>`).join('');
  fReason.innerHTML = '<option value="">全部</option>' + reasons.map(r => `<option>${escapeHtml(r)}</option>`).join('');
  fName.value = prevName; fReason.value = prevReason;
}

// =============== MANAGER TAB ===============
function parseUploadedRows(rows) {
  // rows: array of objects (from SheetJS sheet_to_json) OR array of arrays
  // We'll detect headers; map to standard fields.
  if (!rows.length) return [];
  const sample = rows[0];
  const isObjectRows = !Array.isArray(sample) && typeof sample === 'object';
  const norm = h => String(h || '').trim().toLowerCase().replace(/[\s【】\[\]()（）]/g, '');
  const aliases = {
    name:  ['你的名字','姓名','name','員工','員工姓名'],
    start: ['預假起日','起日','開始日期','start','startdate','from','起'],
    end:   ['預假迄日','迄日','結束日期','end','enddate','to','迄'],
    cat:   ['預假天數','天數','類別','daysrange','category','days'],
    sub:   ['送出時間','申請時間','送出','submittedat','submittime','timestamp'],
    status:['審核結果','結果','狀態','status','result'],
  };
  function findKey(headers, kind) {
    const cands = aliases[kind].map(norm);
    for (const h of headers) if (cands.includes(norm(h))) return h;
    return null;
  }

  let headerRow, dataRows;
  if (isObjectRows) {
    headerRow = Object.keys(sample);
    dataRows = rows;
  } else {
    headerRow = sample.map(String);
    dataRows = rows.slice(1).map(r => Object.fromEntries(headerRow.map((h, i) => [h, r[i]])));
  }
  const kName  = findKey(headerRow, 'name');
  const kStart = findKey(headerRow, 'start');
  const kEnd   = findKey(headerRow, 'end');
  const kCat   = findKey(headerRow, 'cat');
  const kSub   = findKey(headerRow, 'sub');
  if (!kName || !kStart || !kEnd) {
    toast(`找不到必要欄位（姓名/起日/迄日）。表頭：${headerRow.join(', ')}`);
    return [];
  }
  return dataRows.map(r => ({
    id: uid(),
    name: String(r[kName] ?? '').trim(),
    start: toIso(r[kStart]),
    end:   toIso(r[kEnd]),
    daysCat: kCat ? String(r[kCat] ?? '').trim() : '',
    submittedAt: kSub ? String(r[kSub] ?? '') : '',
    decision: 'auto', // auto means follow prediction
    predicted: { cls: 'other', reason: '', conflicts: [] },
  })).filter(e => e.name || e.start || e.end);
}

function handleFile(file) {
  const reader = new FileReader();
  reader.onload = (ev) => {
    try {
      const wb = XLSX.read(ev.target.result, { type: 'array', cellDates: false });
      const ws = wb.Sheets[wb.SheetNames[0]];
      const rows = XLSX.utils.sheet_to_json(ws, { defval: '', raw: true });
      const parsed = parseUploadedRows(rows);
      if (!parsed.length) return;
      state.batch.push(...parsed);
      saveBatch();
      renderManager();
      toast(`已加入 ${parsed.length} 筆申請`);
    } catch (e) {
      console.error(e);
      toast('讀取檔案失敗：' + e.message);
    }
  };
  reader.readAsArrayBuffer(file);
}

function renderBatch() {
  recomputeBatch();
  $('batchCount').textContent = state.batch.length;
  const body = $('batchBody');
  if (!state.batch.length) {
    body.innerHTML = `<tr><td colspan="8" class="empty">尚未上傳任何申請</td></tr>`;
    return;
  }
  body.innerHTML = state.batch.map(e => {
    const days = (e.start && e.end) ? daysInRange(e.start, e.end) : '';
    const eff = e.decision === 'auto' ? e.predicted.cls : e.decision;
    const pillCls = eff === 'pass' ? 'pass' : eff === 'fail' ? 'fail' : 'other';
    const pillLabel = eff === 'pass' ? '通過' : eff === 'fail' ? '未通過' : '—';
    const conflictHtml = e.predicted.conflicts.length
      ? `<div class="conflict-list">已滿日：${e.predicted.conflicts.map(c => `${c.d}(${c.cur})`).join(', ')}</div>`
      : '';
    const predLabel = e.predicted.cls === 'pass' ? '通過'
      : e.predicted.cls === 'fail' ? `未通過 - ${e.predicted.reason}` : '—';
    const predPill = `<span class="pill ${e.predicted.cls === 'pass' ? 'pass' : 'fail'}">${escapeHtml(predLabel)}</span>`;
    const catOptions = ['1-3天','4-10天','>10天'];
    const hasStdCat = !e.daysCat || catOptions.includes(e.daysCat);
    const catSelect = `<select class="edit-cell edit-cat" data-id="${e.id}">
        <option value="" ${!e.daysCat?'selected':''}>未填</option>
        ${catOptions.map(o => `<option value="${escapeHtml(o)}" ${e.daysCat===o?'selected':''}>${escapeHtml(o==='>10天'?'10天以上':o)}</option>`).join('')}
        ${hasStdCat ? '' : `<option value="${escapeHtml(e.daysCat)}" selected>${escapeHtml(e.daysCat)}</option>`}
      </select>`;
    return `<tr>
      <td><input class="edit-cell edit-name" type="text" data-id="${e.id}" value="${escapeHtml(e.name)}" placeholder="姓名" /></td>
      <td><input class="edit-cell edit-date edit-start" type="date" data-id="${e.id}" value="${escapeHtml(e.start || '')}" /></td>
      <td><input class="edit-cell edit-date edit-end" type="date" data-id="${e.id}" value="${escapeHtml(e.end || '')}" /></td>
      <td class="num">${days}</td>
      <td>${catSelect}</td>
      <td>${predPill}${conflictHtml}</td>
      <td>
        <select class="decision" data-id="${e.id}">
          <option value="auto"  ${e.decision==='auto'?'selected':''}>跟隨系統 (${eff==='pass'?'通過':'未通過'})</option>
          <option value="pass"  ${e.decision==='pass'?'selected':''}>強制通過</option>
          <option value="fail"  ${e.decision==='fail'?'selected':''}>強制未通過</option>
        </select>
      </td>
      <td><button class="ghost" data-remove="${e.id}">移除</button></td>
    </tr>`;
  }).join('');

  body.querySelectorAll('select.decision').forEach(sel => {
    sel.onchange = () => {
      const e = state.batch.find(x => x.id === sel.dataset.id);
      if (!e) return;
      e.decision = sel.value;
      saveBatch();
      renderBatch();
    };
  });
  function bindEdit(selector, field, transform) {
    body.querySelectorAll(selector).forEach(el => {
      el.onchange = () => {
        const entry = state.batch.find(x => x.id === el.dataset.id);
        if (!entry) return;
        entry[field] = transform ? transform(el.value) : el.value;
        saveBatch();
        renderBatch();
      };
    });
  }
  bindEdit('input.edit-name',  'name',    v => v.trim());
  bindEdit('input.edit-start', 'start',   v => v || null);
  bindEdit('input.edit-end',   'end',     v => v || null);
  bindEdit('select.edit-cat',  'daysCat');
  body.querySelectorAll('button[data-remove]').forEach(btn => {
    btn.onclick = () => {
      state.batch = state.batch.filter(x => x.id !== btn.dataset.remove);
      saveBatch();
      renderBatch();
    };
  });
  renderDayGrid();
}

function renderDayGrid() {
  // Show day-by-day occupancy for any date touched by the batch.
  const dates = new Set();
  for (const e of state.batch) {
    if (e.start && e.end && e.end >= e.start) {
      for (const d of iterDates(e.start, e.end)) dates.add(d);
    }
  }
  if (!dates.size) { $('dayGrid').innerHTML = '<div class="help">本批尚無有效日期。</div>'; return; }
  const map = buildDayMap(true);
  const sorted = Array.from(dates).sort();
  $('dayGrid').innerHTML = sorted.map(d => {
    const n = map.get(d) || 0;
    const cls = n >= state.quota ? 'full' : (n === state.quota - 1 ? 'tight' : '');
    const dow = ['日','一','二','三','四','五','六'][new Date(d+'T00:00:00').getDay()];
    return `<div class="day-cell ${cls}"><div class="d">${d} (${dow})</div><div class="n">${n} / ${state.quota}</div></div>`;
  }).join('');
}

function commitBatch() {
  if (!state.batch.length) { toast('批次是空的'); return; }
  const stored = loadStored();
  let added = 0;
  for (const e of state.batch) {
    const eff = e.decision === 'auto' ? e.predicted.cls : e.decision;
    if (!e.name || !e.start || !e.end) continue;
    const status = eff === 'pass' ? '通過' : `未通過 - ${e.predicted.reason || '管理員拒絕'}`;
    const rec = {
      '你的名字': e.name,
      '審核結果': status,
      '預假【起日】': null,
      '預假【迄日】': null,
      '送出時間': e.submittedAt || new Date().toISOString().slice(0, 16).replace('T', ' '),
      '預假天數': e.daysCat || '',
      '預假日': null, '特休': null, '時數': null,
      '_start_iso': e.start,
      '_end_iso': e.end,
      '_source': 'manager',
    };
    stored.push(rec);
    added++;
  }
  saveStored(stored);
  state.batch = [];
  saveBatch();
  renderManager();
  renderView();
  toast(`已加入 ${added} 筆紀錄`);
}

function exportBatch() {
  if (!state.batch.length) { toast('批次是空的'); return; }
  const rows = state.batch.map(e => {
    const eff = e.decision === 'auto' ? e.predicted.cls : e.decision;
    return {
      '你的名字': e.name,
      '審核結果': eff === 'pass' ? '通過' : `未通過 - ${e.predicted.reason || '管理員拒絕'}`,
      '預假【起日】': e.start,
      '預假【迄日】': e.end,
      '預假天數': e.daysCat,
      '送出時間': e.submittedAt || '',
    };
  });
  const ws = XLSX.utils.json_to_sheet(rows);
  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, '審核結果');
  XLSX.writeFile(wb, `審核結果_${new Date().toISOString().slice(0,10)}.xlsx`);
}

function renderCommittedRecords() {
  if (!MANAGER_UNLOCKED) return;
  let panel = document.getElementById('committedPanel');
  if (!panel) {
    panel = document.createElement('div');
    panel.className = 'panel';
    panel.id = 'committedPanel';
    panel.innerHTML = `
      <h2>已加入紀錄（手動加入，<span id="committedCount">0</span> 筆）</h2>
      <div class="help">已加入紀錄的手動項目。直接修改欄位即時生效；點刪除可移除單筆。</div>
      <div class="table-wrap" style="border:none;">
        <table>
          <thead><tr>
            <th class="no-sort">姓名</th>
            <th class="no-sort">結果</th>
            <th class="no-sort">拒絕原因</th>
            <th class="no-sort">起日</th>
            <th class="no-sort">迄日</th>
            <th class="no-sort">類別</th>
            <th class="no-sort">操作</th>
          </tr></thead>
          <tbody id="committedBody"></tbody>
        </table>
      </div>`;
    document.getElementById('managerContent').appendChild(panel);
  }

  const stored = loadStored();
  const managerIdx = [];
  stored.forEach((r, i) => { if (r._source === 'manager') managerIdx.push(i); });
  $('committedCount').textContent = managerIdx.length;
  const body = $('committedBody');
  if (!managerIdx.length) {
    body.innerHTML = `<tr><td colspan="7" class="empty">尚未加入任何手動紀錄</td></tr>`;
    return;
  }
  const catOptions = ['1-3天','4-10天','>10天'];
  body.innerHTML = managerIdx.map(i => {
    const r = stored[i];
    const cls = classifyStatus(r['審核結果']);
    const reason = reasonOf(r['審核結果']);
    const cur = r['預假天數'] || '';
    const hasStdCat = !cur || catOptions.includes(cur);
    return `<tr>
      <td><input class="edit-cell c-name" data-i="${i}" value="${escapeHtml(r['你的名字'] || '')}" /></td>
      <td>
        <select class="edit-cell c-status" data-i="${i}">
          <option value="pass" ${cls==='pass'?'selected':''}>通過</option>
          <option value="fail" ${cls==='fail'?'selected':''}>未通過</option>
        </select>
      </td>
      <td><input class="edit-cell c-reason" data-i="${i}" value="${escapeHtml(reason)}" placeholder="${cls==='fail'?'必填':'(僅未通過時使用)'}" ${cls!=='fail'?'disabled':''} /></td>
      <td><input class="edit-cell edit-date c-start" type="date" data-i="${i}" value="${escapeHtml(r._start_iso || '')}" /></td>
      <td><input class="edit-cell edit-date c-end" type="date" data-i="${i}" value="${escapeHtml(r._end_iso || '')}" /></td>
      <td>
        <select class="edit-cell c-cat" data-i="${i}">
          <option value="" ${!cur?'selected':''}>未填</option>
          ${catOptions.map(o => `<option value="${escapeHtml(o)}" ${cur===o?'selected':''}>${escapeHtml(o==='>10天'?'10天以上':o)}</option>`).join('')}
          ${hasStdCat ? '' : `<option value="${escapeHtml(cur)}" selected>${escapeHtml(cur)}</option>`}
        </select>
      </td>
      <td><button class="danger" data-cdel="${i}">刪除</button></td>
    </tr>`;
  }).join('');

  function update(el, fn) {
    const arr = loadStored();
    const r = arr[Number(el.dataset.i)];
    if (!r) return;
    fn(r, el.value);
    saveStored(arr);
    renderView();
    renderCommittedRecords();
    renderBatch();
  }
  body.querySelectorAll('input.c-name').forEach(el => el.onchange = () => update(el, (r, v) => r['你的名字'] = v.trim()));
  body.querySelectorAll('input.c-start').forEach(el => el.onchange = () => update(el, (r, v) => r._start_iso = v || null));
  body.querySelectorAll('input.c-end').forEach(el => el.onchange = () => update(el, (r, v) => r._end_iso = v || null));
  body.querySelectorAll('select.c-cat').forEach(el => el.onchange = () => update(el, (r, v) => r['預假天數'] = v));
  body.querySelectorAll('select.c-status').forEach(el => el.onchange = () => update(el, (r, v) => {
    if (v === 'pass') r['審核結果'] = '通過';
    else r['審核結果'] = '未通過 - ' + (reasonOf(r['審核結果']) || '管理員手動標記');
  }));
  body.querySelectorAll('input.c-reason').forEach(el => el.onchange = () => update(el, (r, v) => {
    if (classifyStatus(r['審核結果']) === 'fail') {
      r['審核結果'] = '未通過 - ' + (v.trim() || '管理員手動標記');
    }
  }));
  body.querySelectorAll('button[data-cdel]').forEach(btn => btn.onclick = () => {
    if (!confirm('確定刪除此筆手動加入紀錄？此動作不會影響原始 Excel。')) return;
    const arr = loadStored();
    arr.splice(Number(btn.dataset.cdel), 1);
    saveStored(arr);
    renderView();
    renderCommittedRecords();
    renderBatch();
    toast('已刪除');
  });
}

function renderManager() {
  $('quota').value = state.quota;
  $('minDays').value = state.minDays;
  $('maxDays').value = state.maxDays;
  $('yearlyPoints').value = state.yearlyPoints;
  $('gateDay').value = state.gateDay;
  const w = roundWindow();
  $('windowHelp').innerHTML = w
    ? `<b>本輪可預約範圍：${w.from} ～ ${w.to}</b>（含）。單筆 ${state.minDays}–${state.maxDays} 天、每日上限 ${state.quota} 人、每人每年 ${state.yearlyPoints} 次核准。`
    : `規則：單筆 ${state.minDays}–${state.maxDays} 天、每日上限 ${state.quota} 人、每人每年 ${state.yearlyPoints} 次核准。<br/>請填入 Gate Day 以啟用「可預約範圍」檢查。`;
  renderBatch();
  renderCommittedRecords();
}

// =============== EDIT MODAL ===============
const editModalState = { handle: null };

function findRecordByHandle(h) {
  if (!h) return null;
  if (h.startsWith('b:')) {
    const idx = Number(h.slice(2));
    return effectiveBaked().find(x => x._baked_idx === idx) || null;
  }
  if (h.startsWith('m:')) {
    return loadStored().find(x => x._id === h.slice(2)) || null;
  }
  return null;
}

function openEditModal(h) {
  const r = findRecordByHandle(h);
  if (!r) { toast('找不到紀錄'); return; }
  editModalState.handle = h;
  const isBaked = h.startsWith('b:');
  $('emMeta').textContent = isBaked ? '原始紀錄（編輯後僅在本機覆寫，不影響原 Excel）' : '手動加入紀錄';
  $('emName').value = r['你的名字'] || '';
  const cls = classifyStatus(r['審核結果']);
  $('emStatus').value = cls === 'pass' ? 'pass' : 'fail';
  $('emReason').value = reasonOf(r['審核結果']);
  $('emStart').value = r._start_iso || '';
  $('emEnd').value = r._end_iso || '';
  $('emCat').value = r['預假天數'] || '';
  $('editModal').classList.remove('hidden');
  setTimeout(() => $('emName').focus(), 0);
}

function closeEditModal() {
  editModalState.handle = null;
  $('editModal').classList.add('hidden');
}

function saveEditModal() {
  const h = editModalState.handle;
  if (!h) return;
  const name = $('emName').value.trim();
  const start = $('emStart').value || null;
  const end = $('emEnd').value || null;
  if (!name) { toast('姓名不可空白'); return; }
  if (start && end && end < start) { toast('迄日早於起日'); return; }
  const status = $('emStatus').value;
  const reason = $('emReason').value.trim();
  const cat = $('emCat').value;
  const statusStr = status === 'pass' ? '通過' : '未通過 - ' + (reason || '管理員手動標記');

  if (h.startsWith('b:')) {
    const idx = Number(h.slice(2));
    const patches = loadBakedPatches();
    patches[idx] = Object.assign({}, patches[idx] || {}, {
      '你的名字': name,
      '審核結果': statusStr,
      '預假天數': cat,
      _start_iso: start,
      _end_iso: end,
    });
    delete patches[idx]._deleted;
    saveBakedPatches(patches);
  } else if (h.startsWith('m:')) {
    const arr = loadStored();
    const r = arr.find(x => x._id === h.slice(2));
    if (!r) { toast('找不到紀錄'); return; }
    r['你的名字'] = name;
    r['審核結果'] = statusStr;
    r['預假天數'] = cat;
    r._start_iso = start;
    r._end_iso = end;
    saveStored(arr);
  }
  closeEditModal();
  renderView();
  if (MANAGER_UNLOCKED) renderManager();
  toast('已儲存');
}

function deleteRecordByHandle(h) {
  const r = findRecordByHandle(h);
  if (!r) return;
  const isBaked = h.startsWith('b:');
  const label = `${r['你的名字'] || '(無姓名)'}（${r._start_iso || '?'} ~ ${r._end_iso || '?'}）`;
  if (!confirm(`確定刪除：${label}？${isBaked ? '\n此動作只在本機覆寫，不會修改原始 Excel；可日後重新整理修補表移除覆寫。' : ''}`)) return;
  if (isBaked) {
    const idx = Number(h.slice(2));
    const patches = loadBakedPatches();
    patches[idx] = Object.assign({}, patches[idx] || {}, { _deleted: true });
    saveBakedPatches(patches);
  } else if (h.startsWith('m:')) {
    const arr = loadStored().filter(x => x._id !== h.slice(2));
    saveStored(arr);
  }
  renderView();
  if (MANAGER_UNLOCKED) renderManager();
  toast('已刪除');
}

function bindEditModal() {
  $('emCancel').onclick = closeEditModal;
  $('emSave').onclick = saveEditModal;
  $('editModal').addEventListener('click', (e) => { if (e.target.id === 'editModal') closeEditModal(); });
  document.addEventListener('keydown', (e) => {
    if (e.key === 'Escape' && !$('editModal').classList.contains('hidden')) closeEditModal();
  });
  $('emStatus').addEventListener('change', () => {
    $('emReason').disabled = $('emStatus').value !== 'fail';
  });
}

// =============== TAB SWITCHING ===============
function showTab(name) {
  document.querySelectorAll('.tab').forEach(t => t.classList.toggle('active', t.dataset.tab === name));
  $('tab-view').classList.toggle('hidden', name !== 'view');
  $('tab-calendar').classList.toggle('hidden', name !== 'calendar');
  $('tab-manager').classList.toggle('hidden', name !== 'manager');
  if (name === 'view') renderView();
  else if (name === 'calendar') renderCalendar();
  else if (MANAGER_UNLOCKED) renderManager();
  else { setTimeout(() => { const i = document.getElementById('pwInput'); if (i) i.focus(); }, 0); }
}

// =============== CALENDAR TAB ===============
const CAL_FULL_THRESHOLD = 2;
const calState = { year: 0, month: 0, selected: '' };

function calOccupancy() {
  const map = new Map();
  for (const r of allRecords()) {
    if (classifyStatus(r['審核結果']) !== 'pass') continue;
    const s = r._start_iso, e = r._end_iso;
    if (!s || !e || e < s) continue;
    for (const d of iterDates(s, e)) {
      if (!map.has(d)) map.set(d, []);
      map.get(d).push({ name: r['你的名字'] || '(無姓名)', start: s, end: e });
    }
  }
  return map;
}

function calLevel(count) {
  if (count <= 0) return 0;
  if (count < CAL_FULL_THRESHOLD) return 1;
  return 2;
}

function renderCalendar() {
  if (!calState.year) initCalendarMonth();
  const occ = calOccupancy();
  const y = calState.year, m = calState.month;
  $('calMonthLabel').textContent = `${y} 年 ${String(m).padStart(2, '0')} 月`;

  const startDow = new Date(y, m - 1, 1).getDay();
  const daysInMonth = new Date(y, m, 0).getDate();

  const heads = ['日','一','二','三','四','五','六'];
  let html = heads.map(h => `<div class="cal-head">${h}</div>`).join('');
  for (let i = 0; i < startDow; i++) html += '<div class="cal-cell empty"></div>';
  for (let d = 1; d <= daysInMonth; d++) {
    const iso = `${y}-${String(m).padStart(2,'0')}-${String(d).padStart(2,'0')}`;
    const list = occ.get(iso) || [];
    const cnt = list.length;
    const lvl = calLevel(cnt);
    const sel = iso === calState.selected ? ' selected' : '';
    const label = cnt === 0 ? '空' : `${cnt} 人`;
    html += `<div class="cal-cell lvl${lvl}${sel}" data-d="${iso}"><div class="num">${d}</div><div class="cnt">${label}</div></div>`;
  }
  const total = startDow + daysInMonth;
  const pad = (7 - (total % 7)) % 7;
  for (let i = 0; i < pad; i++) html += '<div class="cal-cell empty"></div>';

  const grid = $('calGrid');
  grid.innerHTML = html;
  grid.querySelectorAll('.cal-cell[data-d]').forEach(c => {
    c.onclick = () => { calState.selected = c.dataset.d; renderCalendar(); };
  });
  renderCalendarDay(occ);
}

function renderCalendarDay(occ) {
  const iso = calState.selected;
  if (!iso) {
    $('calDayTitle').textContent = '點選日期以查看當日預假人員';
    $('calDayBody').innerHTML = '<div class="help">尚未選擇日期。</div>';
    return;
  }
  const list = (occ || calOccupancy()).get(iso) || [];
  $('calDayTitle').textContent = `${iso}（${list.length} 人預假）`;
  if (!list.length) {
    $('calDayBody').innerHTML = '<div class="help">本日尚無人預假。</div>';
    return;
  }
  const sorted = list.slice().sort((a, b) => a.start.localeCompare(b.start));
  const rows = sorted.map(x =>
    `<tr><td>${escapeHtml(x.name)}</td><td>${x.start}</td><td>${x.end}</td></tr>`
  ).join('');
  $('calDayBody').innerHTML =
    `<table class="cal-day-list"><thead><tr><th>姓名</th><th>起日</th><th>迄日</th></tr></thead><tbody>${rows}</tbody></table>`;
}

function calMove(delta) {
  let { year: y, month: m } = calState;
  m += delta;
  while (m < 1) { m += 12; y -= 1; }
  while (m > 12) { m -= 12; y += 1; }
  calState.year = y; calState.month = m; calState.selected = '';
  renderCalendar();
}

function initCalendarMonth() {
  const today = new Date();
  calState.year = today.getFullYear();
  calState.month = today.getMonth() + 1;
  // If today's month has no records and the baked range is elsewhere, jump to first month with data.
  const hasToday = BAKED.some(r => r._start_iso && r._start_iso.startsWith(`${calState.year}-${String(calState.month).padStart(2,'0')}`));
  if (!hasToday) {
    const dates = BAKED.map(r => r._start_iso).filter(Boolean).sort();
    if (dates.length) {
      const [y, m] = dates[0].split('-');
      calState.year = Number(y); calState.month = Number(m);
    }
  }
}

function bindCalendar() {
  $('calPrev').onclick = () => calMove(-1);
  $('calNext').onclick = () => calMove(1);
  $('calToday').onclick = () => {
    const t = new Date();
    calState.year = t.getFullYear(); calState.month = t.getMonth() + 1; calState.selected = '';
    renderCalendar();
  };
}

// =============== BIND ===============
function bindView() {
  const onChange = () => { state.page = 1; renderView(); };
  $('q').oninput      = (e) => { state.q = e.target.value; onChange(); };
  $('fName').onchange = (e) => { state.name = e.target.value; onChange(); };
  $('fStatus').onchange = (e) => { state.status = e.target.value; onChange(); };
  $('fReason').onchange = (e) => { state.reason = e.target.value; onChange(); };
  $('fFrom').onchange = (e) => { state.from = e.target.value; onChange(); };
  $('fTo').onchange   = (e) => { state.to = e.target.value; onChange(); };
  $('pageSize').onchange = (e) => { state.pageSize = Number(e.target.value); onChange(); };
}

function bindManager() {
  $('quota').onchange = (e) => { state.quota = Math.max(1, Number(e.target.value) || 1); renderManager(); };
  $('minDays').onchange = (e) => { state.minDays = Math.max(1, Number(e.target.value) || 1); renderManager(); };
  $('maxDays').onchange = (e) => { state.maxDays = Math.max(1, Number(e.target.value) || 1); renderManager(); };
  $('yearlyPoints').onchange = (e) => { state.yearlyPoints = Math.max(1, Number(e.target.value) || 1); renderManager(); };
  $('gateDay').onchange = (e) => { state.gateDay = e.target.value; renderManager(); };

  const drop = $('drop'), input = $('fileInput');
  drop.onclick = () => input.click();
  input.onchange = () => { if (input.files[0]) { handleFile(input.files[0]); input.value = ''; } };
  ['dragenter','dragover'].forEach(ev => drop.addEventListener(ev, e => { e.preventDefault(); drop.classList.add('dragover'); }));
  ['dragleave','drop'].forEach(ev => drop.addEventListener(ev, e => { e.preventDefault(); drop.classList.remove('dragover'); }));
  drop.addEventListener('drop', e => {
    const f = e.dataTransfer.files[0]; if (f) handleFile(f);
  });

  $('mAdd').onclick = () => {
    const name = $('mName').value.trim();
    const start = $('mStart').value;
    const end = $('mEnd').value;
    const daysCat = $('mDays').value;
    if (!name || !start || !end) { toast('請填寫姓名、起日、迄日'); return; }
    state.batch.push({ id: uid(), name, start, end, daysCat, submittedAt: '', decision: 'auto', predicted: { cls:'other', reason:'', conflicts:[] } });
    saveBatch();
    $('mName').value = ''; $('mStart').value = ''; $('mEnd').value = '';
    renderBatch();
    toast('已加入');
  };

  $('commit').onclick = () => {
    if (!confirm(`確定將 ${state.batch.length} 筆最終決定加入紀錄？此動作會更新檢視紀錄分頁。`)) return;
    commitBatch();
  };
  $('clearBatch').onclick = () => {
    if (!state.batch.length) return;
    if (!confirm(`清空目前 ${state.batch.length} 筆批次申請？`)) return;
    state.batch = []; saveBatch(); renderBatch();
  };
  $('exportBatch').onclick = exportBatch;
  $('resetStored').onclick = () => {
    const n = loadStored().length;
    if (!n) { toast('沒有已儲存的新增紀錄'); return; }
    if (!confirm(`確定移除全部 ${n} 筆已儲存的新增紀錄？此動作不會影響原始 Excel。`)) return;
    saveStored([]); renderManager(); renderView(); toast('已移除');
  };
}

document.querySelectorAll('.tab').forEach(t => t.onclick = () => showTab(t.dataset.tab));

// =============== INIT ===============
function showRange() {
  const dates = BAKED.map(r => r._start_iso).filter(Boolean).sort();
  if (dates.length) $('rangeMeta').textContent = `預假起日範圍：${dates[0]} ～ ${dates[dates.length - 1]}（原始資料 ${BAKED.length} 筆）`;
}

// =============== MANAGER UNLOCK (PBKDF2 + AES-GCM) ===============
const ENC = __ENC__;  // {salt, iv, ct, iters} all base64 (except iters)
let MANAGER_UNLOCKED = false;

function b64ToBytes(s) {
  const bin = atob(s); const out = new Uint8Array(bin.length);
  for (let i = 0; i < bin.length; i++) out[i] = bin.charCodeAt(i);
  return out;
}

async function deriveKey(password) {
  const enc = new TextEncoder();
  const baseKey = await crypto.subtle.importKey('raw', enc.encode(password),
    { name: 'PBKDF2' }, false, ['deriveKey']);
  return crypto.subtle.deriveKey(
    { name: 'PBKDF2', salt: b64ToBytes(ENC.salt), iterations: ENC.iters, hash: 'SHA-256' },
    baseKey,
    { name: 'AES-GCM', length: 256 },
    false, ['decrypt']);
}

async function unlockManager(password) {
  const key = await deriveKey(password);
  const plain = await crypto.subtle.decrypt(
    { name: 'AES-GCM', iv: b64ToBytes(ENC.iv) }, key, b64ToBytes(ENC.ct));
  const html = new TextDecoder().decode(plain);
  document.getElementById('managerContent').innerHTML = html;
  document.getElementById('lockScreen').classList.add('hidden');
  MANAGER_UNLOCKED = true;
  bindManager();
  renderManager();
  renderView();
}

document.getElementById('unlockForm').addEventListener('submit', async (e) => {
  e.preventDefault();
  const pw = document.getElementById('pwInput').value;
  const err = document.getElementById('unlockErr');
  err.textContent = '';
  if (!pw) { err.textContent = '請輸入密碼'; return; }
  try {
    await unlockManager(pw);
    document.getElementById('pwInput').value = '';
  } catch {
    err.textContent = '密碼錯誤';
  }
});

state.batch = loadBatch();
bindView();
bindCalendar();
bindEditModal();
showRange();
renderView();
</script>
</body>
</html>
"""


MANAGER_HTML_BLOCK = """\
<div class="panel">
  <h2>審核規則</h2>
  <div class="controls">
    <label>每日通過上限人數
      <input id="quota" type="number" min="1" value="2" />
    </label>
    <label>單筆預假最少天數
      <input id="minDays" type="number" min="1" value="4" />
    </label>
    <label>單筆預假最多天數
      <input id="maxDays" type="number" min="1" value="10" />
    </label>
    <label>個人年度核准次數上限（點數）
      <input id="yearlyPoints" type="number" min="1" value="12" />
    </label>
    <label>本輪 Gate Day（首週六）
      <input id="gateDay" type="date" />
    </label>
  </div>
  <div class="help" id="windowHelp">
    規則：單筆 4–10 天、每日最多 2 人、每人每年 12 次核准（每筆通過扣 1 點，依起日年份計算），
    預假日須落在 <b>Gate Day</b> 至 <b>(Gate Day + 7 個月) 之次月首個週日</b> 之間。
    留空 Gate Day 則略過範圍檢查。
  </div>
</div>

<div class="panel">
  <h2>上傳新申請</h2>
  <div id="drop" class="drop">
    將 .xlsx 或 .csv 拖曳到此處，或 <strong>點擊選擇檔案</strong>
    <input id="fileInput" type="file" accept=".xlsx,.xls,.csv" hidden />
    <div class="help">必要欄位：你的名字、預假【起日】、預假【迄日】。可選欄位：預假天數、送出時間。日期可為 YYYY-MM-DD、Excel 序號、或一般日期格式。</div>
  </div>

  <details>
    <summary class="help" style="cursor:pointer; margin-top:10px;">或手動新增單筆</summary>
    <div class="controls" style="margin-top:10px;">
      <label>姓名 <input id="mName" type="text" /></label>
      <label>起日 <input id="mStart" type="date" /></label>
      <label>迄日 <input id="mEnd" type="date" /></label>
      <label>類別
        <select id="mDays">
          <option value="">未填</option>
          <option value="1-3天">1-3天</option>
          <option value="4-10天" selected>4-10天</option>
          <option value=">10天">10天以上</option>
        </select>
      </label>
      <label>&nbsp;<button class="primary" id="mAdd">新增到批次</button></label>
    </div>
  </details>
</div>

<div class="panel">
  <h2>本批新申請（<span id="batchCount">0</span> 筆）</h2>
  <div class="table-wrap" style="border:none;">
    <table>
      <thead><tr>
        <th class="no-sort">姓名</th>
        <th class="no-sort">起日</th>
        <th class="no-sort">迄日</th>
        <th class="no-sort">天數</th>
        <th class="no-sort">類別</th>
        <th class="no-sort">系統判定</th>
        <th class="no-sort">最終決定</th>
        <th class="no-sort">操作</th>
      </tr></thead>
      <tbody id="batchBody"><tr><td colspan="8" class="empty">尚未上傳任何申請</td></tr></tbody>
    </table>
  </div>
  <div class="actions-row">
    <button class="primary" id="commit">將最終決定加入紀錄</button>
    <button class="ghost" id="exportBatch">匯出本批 xlsx</button>
    <button class="ghost" id="clearBatch">清空批次</button>
    <div class="spacer"></div>
    <button class="danger" id="resetStored">移除所有已儲存的新增紀錄</button>
  </div>

  <details class="day-detail" id="dayDetailWrap">
    <summary>顯示本批涉及日期的佔用狀況</summary>
    <div class="day-grid" id="dayGrid"></div>
  </details>
</div>
"""


def main():
    headers, records = load_records()
    data = to_jsonable(records)
    enc = encrypt_manager_block(MANAGER_HTML_BLOCK, MANAGER_PASSWORD)
    html = (HTML_TEMPLATE
            .replace("__DATA__", json.dumps(data, ensure_ascii=False))
            .replace("__HEADERS__", json.dumps(headers, ensure_ascii=False))
            .replace("__ENC__", json.dumps(enc)))
    OUT.write_text(html, encoding="utf-8")
    print(f"Wrote {OUT}  ({OUT.stat().st_size:,} bytes,  {len(data)} records,"
          f"  manager block: {len(MANAGER_HTML_BLOCK)} → ct {len(enc['ct'])} chars b64)")


if __name__ == "__main__":
    main()
